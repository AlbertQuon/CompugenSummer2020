# Address.py
# Scans through an excel file, finds addresses, applies rules of consistency, flags addresses that aren't consistent
# Albert Quon
# Created: 2020/05/06
# Last modified: 2020/08/25

# IMPORTS ###
# EXCEL
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle, PatternFill
# GUI
import PySimpleGUI as gui
# I/O and DEBUGGING
from timeit import default_timer as timer
import os.path
import json
import traceback


def scan(sheet):
    """
    # go through all the addresses, check the province, flag if not consistent
    :param sheet: the spreadsheet
    :return: all addresses
    """
    addresses = []
    provCol = findCol("Province", sheet["1"])
    scanCol = findCol("AddressLine1", sheet["1"])
    scanColB = findCol("AddressLine2", sheet["1"])

    # DATA VALIDATION - check if the address is valid itself
    # structure the address and its ext info to be readable by the program
    # then standardize it through the rules and validate

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  #start below the headers then up to sheet.max_row
        address = structureAddress(scanCol, provCol, row) # structure the string to an address object

        # special case for "FERME PHYSIQUE" as they can be seen as a valid address and should be left alone
        if "FERME PHYSIQUE" not in address.street and not "INVALID" in address.flag.address:
            extAddress = structureAddress(scanColB, provCol, row)
            if "INVALID" not in extAddress.flag.address:
                validate(extAddress)
            validate(address)

            if len(extAddress.original) != 0 and extAddress.original is not None:
                if len(address.extra) == 0:
                    address.extra = " ".join(extAddress.original)
                elif address.extra != " ".join(extAddress.original):
                    address.extra = " ".join(extAddress.original) + ", " + address.extra
            if len(address.extra) > 40: # attempt to reduce character length
                address.extra = trimExtInfo(address.extra)

        elif "INVALID" in address.flag.address: # invalid addresses
            extAddress = structureAddress(scanColB, provCol, row)

            if len(extAddress.original) != 0:
                if len(address.extra) == 0:
                    address.extra = " ".join(extAddress.original)
                elif address.extra != " ".join(extAddress.original):
                    address.extra = " ".join(extAddress.original) + ", " + address.extra
            if len(address.extra) > 40: # attempt to reduce character length
                address.extra = trimExtInfo(address.extra)

        addresses.append(address)

    return addresses


def findCol(column, sheet):
    """
    Finds the column in a spreadsheet with a given name
    :param column: Column name to be found
    :param sheet: Excel worksheet
    :return: Column of the name given
    """
    for cell in sheet:
        if cell.value == column:
            col = sheet.index(cell)
            return col
    return -1


def structureAddress(scanCol, provCol, row):
    """
    Formats the address into an address object. Rejects addresses that cannot be interpreted or have special cases.
    :param row: Row to scan
    :param scanCol: Column of addresses
    :param provCol: Column of provinces
    :return: Address object
    """

    address = str(row[scanCol].value).rstrip().replace("None", "").split()

    combine = "".join(address)
    junction = ("/" in combine and not (any(len(word) < 2 for word in combine.split("/")))) or (
                "JUNCTION" in address) or ("JUNC" in address) or\
               ("&" in combine and not any(len(word) < 2 for word in "".join(address).split("&"))) or\
               "CORNEROF" in combine or "AND" in address # 'AND' in address may be unreliable
    directions = "KM" in address
    #gps = ("N" in address and "W" in address) or ("LONG" in address and "LAT" in address)

    if len(address) < 2 or junction or directions:  # if it is not a valid address; cannot be one 'word' only or be a intersection
        if len(address) > 0:
            if not address[0].isalnum() and len(address[0]) == 1:
                address.pop(0)
        temp = Address(address)
        temp.flag.addAddrFlag("INVALID")
        return temp

    else:  # at least a number and word
        newAddress = Address(address)

        prov = row[provCol].value.strip()
        newAddress.french = prov == "QC"
        address = newAddress.original
        # special case
        if "FERME" in address and "PHYSIQUE" in address:
            newAddress.street = " ".join(address)
            return newAddress

        # Clean the address, make it easy to read ###################
        # BREAK APART COMMAS, BREAK APART PERIODS
        # DETACH DASHES FROM ORDINAL NUMBERS
        i = 0
        length = len(address)

        # address is cleaned of punctuation that can lead to misinterpretation of each section of the address
        while i < length:
            if "," in address[i] and len(address[i].split(",")) > 2 and len(address[i]) > 3 and not any(
                    len(w) <= 1 for w in address[i].split(",")):
                array = address[i].split(",")
                address.pop(i)
                for j in range(len(array)):
                    address.insert(i + j, array[j])
                length += len(array) - 1

            if "." in address[i] and len(address[i].split(".")) > 1 and len(address[i]) > 3 and not any(
                    len(w) <= 1 for w in address[i].split(".")):
                array = address[i].split(".")
                address.pop(i)
                for j in range(len(array)):
                    address.insert(i + j, array[j])
                length += len(array) - 1

            if "\\" in address[i] and len(address[i].split("\\")) > 1 and len(address[i]) > 3 and not any(
                    len(w) <= 1 and w.isalpha() for w in address[i].split("\\")):
                array = address[i].split("\\")
                address.pop(i)
                for j in range(len(array)):
                    address.insert(i + j, array[j])
                length += len(array) - 1

            if "/" in address[i] and len(address[i].split("/")) > 1 and len(address[i]) > 3 and not any(
                    len(w) <= 1 and w.isalpha() for w in address[i].split("/")):
                array = address[i].split("/")
                address.pop(i)
                for j in range(len(array)):
                    address.insert(i + j, array[j])
                length += len(array) - 1

            if "-" in address[i] and len(address[i].split('-')) == 2 and not\
                    any(not checkNumbers(w) for w in address[i].split('-')) and (
                    any(checkOrdinal(w) for w in address[i].split("-")) or i > 0):
                array = address[i].split("-")
                address.pop(i)
                for j in range(len(array)):
                    address.insert(i + j, array[j])
                length += len(array) - 1

            if address[i] == "-" and len(address) - 1 > i > 0:
                noSym = "".join(letter for letter in address[i - 1] + address[i + 1] if letter.isalnum())
                if not noSym.isdigit() or not noSym.isalpha():
                    address.pop(i)
                length -= 1

            i += 1

        leftBrac = rightBrac = -1
        extra = number = street = direction = suffixNumber = ""

        # check if brackets exist in the address, and add it to the ext info
        for i in range(len(address)):
            if "(" in address[i] or "[" in address[i]:
                leftBrac = i
            if (")" in address[i] or "]" in address[i]) and leftBrac != -1:
                rightBrac = i
        if leftBrac != rightBrac != -1:
            for i in range(leftBrac, rightBrac + 1):
                extra = " ".join([extra, address[leftBrac]])
                address.pop(leftBrac)
        if not directions: # after symbols are seperated, check if it can be interpreted as directions
            for i in range(len(address)):
                if i < len(address) - 1:
                    if (address[i] == "MILE" or address[i] == "KM" or address[i] == "MILES") and directionFactor(
                            address[i + 1]) and not directions:
                        directions = True

        if len(address) == 0 or "AND" in address or (address[0] == "NE" and "".join(c for c in address[1] if c.isalnum()))\
                or directions:
            temp = Address(address)
            temp.flag.addAddrFlag("INVALID")
            return temp

        # check if the first part of the address is a symbol
        if not address[0].isalnum() and not (
                any(letter.isalpha() for letter in address[0]) or any(num.isdigit() for num in address[0])):
            address.pop(0)

        # Attach dashes and symbols
        first = [address[0]]
        for i in range(1, len(address)):
            if (len(address[i]) == 1 and not address[i].isalnum()) or (any(first[-1][-1] == sym for sym in ["-"])):
                first[-1] += address[i]
            else:
                first.append(address[i])

        # make sure the address is upper case to avoid comparison inconsistencies
        for i in range(len(first)):
            if not first[i].isupper() and any(letter.isalpha() for letter in first[i]):
                newAddress.flag.addAddrFlag("FORMAT")
                first[i] = first[i].upper()

        address = first
        newAddress.original = address

        suffixFactors = []
        fullSuffixes = []
        extFactors = []
        dirFactors = []
        wordCounter = 0
        #************* Assign the direction, street, suffix, and extra info ********************************************
        if newAddress.french:  # determines if more emphasis is put onto the first few words or the last
            frenchFactor = -1
        else:
            frenchFactor = 1

        for i in range(len(address)):  # go through parts of the address, determine factors
            # assign factors based on the location of the part of the address and its characteristics
            numberFactor = sum(c.isnumeric() for c in address[i])
            wordFactor = sum(c.isalpha() for c in address[i])

            # ordinal numbers are not considered numbers
            if checkOrdinal(address[i]):
                numberFactor = 0

            if numberFactor == wordFactor == 0:  # character is a symbol
                suffFactor = extFactor = 0
                dirFactor = False
                if address[i - 1] in number:
                    number += address[i]
                elif address[i - 1] in street:
                    street += address[i]
                else:
                    if len(number) == 0:
                        number = address[i]
                    else:
                        number += address[i]
            else:
                if numberFactor > wordFactor:  # if the current word is a number
                    suffFactor = extFactor = 0
                    dirFactor = False
                    fullSuffixes.append("")

                    if i > 0:  # if the number is not at the beginning of the address
                        if any(word in address[i - 1] for word in
                               ["HWY", "HIGHWAY", "ROUTE", "RTE", "RR", "R.R.", "PTH"]) or address[i - 1] == "NO." or\
                                address[i - 1] == "NO":
                            extFactor = 0
                        elif extFactors[-1] > 1.4 and not removeSymbols(address[i - 1]).isdigit():
                            if i > 1:
                                if not checkOrdinal(address[i - 2]):
                                    extFactor = 1.3
                            else:
                                extFactor = 1.3

                else:  # the current word is a word
                    # determine if its a direction, belongs in an address, or is external info
                    suffixPair = calcSuffixFactor(address[i])
                    suffFactor = suffixPair[0] * 1.15 ** (wordCounter * frenchFactor)
                    dirFactor = directionFactor(address[i])
                    extFactor = calcExtFactor(address[i]) * 1.15 ** wordCounter
                    fullSuffixes.append(suffixPair[1])  # keep note of the actual suffix assigned
                    wordCounter += 1
                    if address[i] == "MAIN":
                        if i > 0:
                            if address[i - 1] == "STATION" or address[i - 1] == "STN":
                                extFactor = extFactors[i - 1]
                        if i < len(address) - 1:
                            if address[i + 1] == "STATION" or address[i + 1] == "STN":
                                extFactor = 2
                    if (removeSymbols(address[i]) == "ST" or removeSymbols(address[i]) == "STE") and (
                            wordCounter == 1 or suffixFactors[-1] >= 1) and i != len(address)-1:
                        suffFactor = extFactor = 0

            extFactors.append(extFactor)
            suffixFactors.append(suffFactor)
            dirFactors.append(dirFactor)
        #print(address)
        #print(extFactors)
        #print(suffixFactors)
        #print(dirFactors)

        # determine the suffix
        suffIndex = suffixFactors.index(max(suffixFactors))

        if dirFactors[suffIndex]:
            suffixFactors[suffIndex] = 0
        suffIndex = suffixFactors.index(max(suffixFactors))
        # assign the suffix
        if suffixFactors[suffIndex] >= 1:
            suffix = address[suffIndex]
            extFactors[suffIndex] = 0
            altSuffix = fullSuffixes[suffIndex]
            # some suffixes are translated between french to english and vice versa, this is to prevent confusion
            if any(altSuffix in frSuffix for frSuffix in ["RUE", "PROMENADE", "CHEMIN"]):
                newAddress.french = True
            if any(altSuffix in engSuffix for engSuffix in ["STREET", "DRIVE", "ROAD", "SIDEROAD", "CRESCENT"]):
                newAddress.french = False
        else:
            suffix = altSuffix = ""
            suffIndex = -1

        # assign street, name, extra info, and, if possible, direction
        for i in range(len(address)):
            numberFactor = sum(c.isdigit() for c in address[i])
            wordFactor = sum(c.isalpha() for c in address[i])
            if checkOrdinal(address[i]): # ordinal numbers are not numbers, thus, should be in the street or ext info
                if i < len(address) - 1:
                    if extFactors[i + 1] > 1.3:
                        extra = " ".join([extra, address[i]])
                    else:
                        street = " ".join([street, address[i]])
                else:
                    street = " ".join([street, address[i]])

            elif numberFactor > wordFactor: # join numbers with highways and rural roads
                if extFactors[i] >= 1.1:
                    extra = " ".join([extra, address[i]])
                elif i > 0:
                    if suffIndex + 1 == i and (not newAddress.french or altSuffix == "ROUTE") and suffIndex != -1:
                        suffixNumber = address[i]
                    elif i > suffIndex != -1:
                        extra = " ".join([extra, address[i]])
                    elif any(word in address[i - 1] for word in
                             ["HWY", "HIGHWAY", "ROUTE", "RTE", "RR", "R.R.", "PTH"]) or address[i - 1] == "NO." or\
                            address[i - 1] == "NO":
                        street = " ".join([street, address[i]])
                        suffIndex = i
                    else:
                        number = " ".join([number, address[i]])
                else:
                    number = " ".join([number, address[i]])
            else:
                if dirFactors[i]:
                    noSym = "".join(letter for letter in address[i] if letter.isalpha())
                    if i < suffIndex != -1 or suffIndex == -1:
                        if len(noSym) == 1:
                            multiDirect = {"N":["NORTH", "NORD"], "E":["EAST", "EST"], "S":["SOUTH", "SUD"], "W":"WEST",
                                           "O":"OUEST"}
                            for letter in multiDirect:
                                if noSym == letter:
                                    if letter == "W" or letter == "O":
                                        address[i] = multiDirect[letter]
                                    elif newAddress.french:
                                        address[i] = multiDirect[letter][1]
                                    else:
                                        address[i] = multiDirect[letter][0]

                        street = " ".join([street, address[i]])

                    else:
                        if len(direction) == 2 or len(direction) > 5:
                            extra = " ".join([extra, address[i]])
                        elif i < len(address) - 1:
                            if "TOWER" == address[i + 1] or "MALL" == address[i + 1]:
                                extra = " ".join([extra, address[i]])
                        if address[i] not in extra:
                            if i > 0:
                                if dirFactors[i - 1] or (len(direction) < 2 or len(direction) <= 5):
                                    direction = "".join([direction, noSym])
                                else:
                                    extra = " ".join([extra, address[i]])
                            else:
                                direction = "".join([direction, noSym])
                else:
                    exception = any(
                        address[i] == word for word in ["STATION", "PORT", "DOCK", "PORTE", "TOWER"]) and i < suffIndex != 0
                    # exception exists in case the street name is mistaken as ext info if it is before the suffix
                    if (extFactors[i] >= 1.1 or (
                            i > suffIndex != -1 and not newAddress.french and altSuffix != "PLACE")) and not exception:
                        extra = " ".join([extra, address[i]])
                    elif i != suffIndex:
                        street = " ".join([street, address[i]])

        street = street.strip()
        extra = extra.strip()
        number = number.strip()

        # special case
        if street == "ST":
            street = street + "||STREET"

        # suffix is actually the street
        if len(street) == 0 and len(suffix) != 0:
            street = suffix + "||" + altSuffix
            suffix = altSuffix = ""
        joined = "".join(address)

        # special case
        if joined.find(suffix) != -1 and joined.find(street) != -1 and len(suffix) > 0 and len(street) > 0:
            if joined.find(suffix) < joined.find(street) and "PLACE" == altSuffix:
                newAddress.french = True

        newAddress.ordinal = any(checkOrdinal(word) for word in street.split())
        newAddress.street = street
        newAddress.original = address
        newAddress.number = number
        newAddress.suffix = suffix
        newAddress.suffixNumber = suffixNumber
        newAddress.direction = direction
        newAddress.extra = extra.rstrip()
        newAddress.altSuffix = altSuffix

        return newAddress


def calcSuffixFactor(word):
    """
    Determines how likely a word is to be a suffix
    :param word: Word to be added
    :return: Highest factor with the suffix that has the highest factor
    """

    suffixFactor = 0
    matchedSuffix = ""
    if ("," not in word and "-" not in word and "." not in word) and not word.isalnum():
        # symbols that aren't used to abbreviate
        return (0, "")

    word = removeSymbols(word)

    # if the word is only a letter, it is impossible to determine if it represents a suffix
    if len(word) <= 1:
        return (0, "")


    for suffix in suffixes:
        i = j = 0
        simFactor = 0
        while i < len(word) and j < len(suffix):
            if word[i] == suffix[j] or word[i] not in suffix:
                if word[i] == suffix[j]:
                    simFactor += suffixes[suffix][j]
                if word[i] not in suffix:
                    simFactor -= max(suffixes[suffix])*2
                #print(word[i], simFactor)
                i += 1


            j += 1

        prev = suffixFactor
        #suffixFactor = max(simFactor, consecFactor, initFactor, suffixFactor)
        suffixFactor = max(simFactor, suffixFactor)
        if suffixFactor != prev:
            matchedSuffix = suffix

    return (suffixFactor, matchedSuffix)


def calcExtFactor(word):
    """
    Determines how likely a word is external info for an address
    :param word: The word
    :return: The highest external factor found
    """
    # check if it's a case of 1st, 2nd, 3rd.
    if len(word) >= 2:
        if checkOrdinal(word):
            return 0.92  # was previously 0.9
        if word == "ST":
            return 0
    else:
        if word.isalpha():
            return 1.05
    symFactor = 0
    for symbol in ",!@#$%&^*()-_+=/:[]0123456789.":
        if symbol in word:
            if symbol == "-" or symbol == "*":  #symbol == "." or symbol == "-"
                if symbol == "*" and word.count(symbol) > 2:
                    symFactor += 1.1 * (word.count(symbol) - 2)
                symFactor += 0.5
            elif symbol == ".":
                if symFactor > 0:
                    if word.count(symbol) >= 2:
                        symFactor += 0.5 * (word.count(symbol))
                    else:
                        symFactor += 0.25
                else:
                    symFactor += 0.5
            elif symbol == ",":
                if len(word) > 5:
                    symFactor += 0.9 / len(word)
                else:
                    symFactor += 0.9
            elif symbol == "&":
                symFactor += 0.5
            else:
                symFactor += 1.1
    if symFactor > 0.5:
        return symFactor

    extFactor = 0
    word = removeSymbols(word)

    if word == "LA" or word == "DE" or len(word) == 0 or word == "OF" or word == "ST": # outliers
        return 0
    if word == "GD":
        return 1.1
    if word == "RR":
        return 0.8

    for i in range(len(extras)):
        k = j = 0
        simFactor = consecFactor = 0
        foreign = False
        while j < len(word) and k < len(extras[i]) and not foreign:
            if word[j] == extras[i][k]:
                if j == k and len(word) >= 3:
                    consecFactor += 1.3 / len(word)
                simFactor += 0.8 / len(word) + 1 / (2 * len(extras[i]))
                j += 1

            k += 1

        if not set(word).issubset(set(extras[i])) or word[0] != extras[i][0]:
            simFactor = consecFactor = 0

        if extras[i] == "SECTION":
            if "C" not in word:
                simFactor = consecFactor = 0
        if extras[i] == "MEZZANINE":
            if "Z" not in word:
                simFactor = consecFactor = 0
        if extras[i] == "PORTE" or extras[i] == "PORT":
            if "T" not in word:
                simFactor = consecFactor = 0

        extFactor = max(simFactor, consecFactor, symFactor, extFactor)


    return extFactor


def directionFactor(word):
    """
    Determines if the word can be a direction
    :param word: The word to be determined
    :return: Boolean value indicating if direction or not
    """
    word = "".join(letter for letter in word if letter.isalpha())

    for direct in ["NORTH", "WEST", "EAST", "SOUTH", "NORD", "OUEST", "EST", "SUD", "SOUTHEAST",
                   "SOUTHWEST", "NORTHEAST", "NORTHWEST", "NORDEST", "NORDOUEST",
                   "SUDEST", "SUDOUEST"]:
        if direct == word or word == direct[0]:
            return True
    for direct in ["N", "S"]:
        for subDirect in ["E", "W"]:
            if direct + subDirect == word:
                return True
    return direct == "O"


def findPOFactor(street):
    """
    Finds an instance of "PO BOX, CP, or CASE POSTALE" in a string
    :param street: The Street name
    :return: Boolean value indicating if PO BOX or not
    """
    PO = "POBOX"
    CP = "CP"
    CasePost = "CASE POSTALE"
    PObag = "POBAG"
    street = street.replace(",", "").replace(".", "").replace(" ", "")
    word = ""
    extra = ""

    for letter in street:
        if letter.isalpha():
            word += letter
        else:
            extra += letter

    if extra.isdigit():
        return PO in word or (CP == word) or CasePost in word or PObag in word or word == "BOX"
    return False


def removeSymbols(word, exception="", remove=""):
    """
    Removes symbols from a word
    :param word: The word to be cleaned
    :param exception: Any characters to be excepted
    :param remove: Any characters that should be removed
    :return: Cleaned word
    """
    newWord = ""
    for letter in word:
        if len(remove) > 0:
            if letter.isalnum() or letter not in remove:
                newWord += letter
        elif letter.isalnum() or letter in "- '" or letter in exception:
            newWord += letter
    return newWord


def checkNumbers(word):
    """
    Checks if any digits exist in a string
    :param word: The string
    :return: Boolean value if digit exists
    """
    return any(c.isdigit() for c in word)


def checkOrdinal(word):
    """
    Checks if a string can be an ordinal
    :param word: The string
    :return: Boolean value if string is an ordinal number
    """
    if not checkNumbers(word) or word.isdigit() or ("-" in word and word.strip("-") == word):
        return False

    word = removeSymbols(removeSymbols(word, "", "' -"))
    letters = ""
    for char in word:
        if not char.isdigit():
            letters += char

    if len(letters) == 1:
        return letters == "E"
    elif len(letters) == 2:
        return letters == "TH" or letters == "ER" or letters == "RE" or letters == "ND" or letters == "RD" or letters == "ST"
    elif len(letters) > 2:
        return "ERE" in letters or letters == "IER" or "ME" in letters or "ÈRE" in letters
    else:
        return False


def validate(address):
    """
    Validate a given address based on the rules given and flag any inconsistencies
    :param address: Address object
    """

    newDir = ""
    newStreet = ""
    newSuffix = ""

    # category stores the errors for each location
    # NUMBER #############################################

    if removeSymbols(address.number) != address.number:
        address.flag.addNumFlag("SYM")
        address.number = removeSymbols(address.number)

    if len(address.number) != 0:
        if address.number[-1] == "-" or address.number[0] == "-":
            address.number = address.number.strip("-")
    else:
        address.flag.addNumFlag("UNDEFINED")

    # DIRECTION ##########################################

    if not address.direction.isupper() and len(address.direction) > 0:
        address.flag.addDirFlag("FORMAT")

    if "." in address.direction or removeSymbols(removeSymbols(address.direction, "", "' -")) != address.direction:
        address.flag.addDirFlag("SYM")
        address.direction = removeSymbols(removeSymbols(address.direction, "", "' -"))

    if address.direction in ["NORTH", "WEST", "EAST", "SOUTH", "NORD", "OUEST", "EST", "SUD", "SOUTHEAST",
                             "SOUTHWEST", "NORTHEAST", "NORTHWEST", "NORDEST", "NORDOUEST", "SUDEST", "SUDOUEST"]:
        address.flag.addDirFlag("LEN")

    if "LEN" in address.flag.direction:
        if "NORTH" in address.direction or "NORD" in address.direction:
            newDir += "N"
        elif "SOUTH" in address.direction or "SUD" in address.direction:
            newDir += "S"
        if "WEST" in address.direction:
            newDir += "W"
        elif "OUEST" in address.direction:
            newDir += "O"
        elif "EST" in address.direction or "EAST" in address.direction:
            newDir += "E"
    else:
        for letter in address.direction:
            if letter.isalpha():
                newDir += letter

    address.direction = newDir
    # STREET #############################################

    isPO = findPOFactor(address.extra) or findPOFactor(address.street)
    name = address.street.split()

    if "" in name:
        address.flag.addStrFlag("FORMAT")

    for symbol in ",.!#$%^&*()[]<>/~;=_+–":
        for word in name:
            if symbol in word:
                address.flag.addStrFlag("SYM")
                name[name.index(word)] = removeSymbols(word)

    if len(name) >= 2:

        for word in name:
            stFlag = steFlag = False
            for saint in ["ST", "STE", "SAINTE", "SAINT"]:
                if saint in word and len(saint) > 3:
                    address.flag.addStrFlag("ST/STE")
                    if "-" in word:
                        saintWord = word[:word.index("-")]
                    else:
                        saintWord = ""
                    if not stFlag and (word == "SAINT" or saintWord == "SAINT"):
                        stFlag = saint == "SAINT"
                    if not steFlag and (word == "SAINTE" or saintWord == "SAINTE"):
                        steFlag = saint == "SAINTE"
                elif len(saint) <= 3 and saint in word:
                    if "-" in word:
                        saintWord = word[:word.index("-")]
                    else:
                        saintWord = ""
                    if not stFlag and (word == "ST" or saintWord == "ST"):
                        stFlag = saint == "ST"
                    if not steFlag and (word == "STE" or saintWord == "STE"):
                        steFlag = saint == "STE"

            if stFlag:
                newStreet += "ST. "
                if "-" in word:
                    newStreet += word.split("-")[-1] + " "
            elif steFlag:
                newStreet += "STE. "
                if "-" in word:
                    newStreet += word.split("-")[-1] + " "
            else:
                newStreet += word + " "

    else:
        # CASE that the address is a PO BOX
        if isPO:
            if address.external:
                address.flag.addStrFlag("EXCESS")
            elif "." in address.extra and len(address.street) == 0:
                address.flag.addStrFlag("SYM")
                address.extra = removeSymbols(address.extra)

        if len(address.street) == 0 and not isPO:
            address.flag.addStrFlag("UNDEFINED")

        if "-" in address.street:
            nameB = address.street.split("-")

            for word in nameB:
                stFlag = steFlag = False
                for saint in ["ST", "STE", "SAINTE", "SAINT"]:
                    if saint in word and len(saint) > 3:
                        address.flag.addStrFlag("ST/STE")
                        if not stFlag and word == "SAINT":
                            stFlag = saint == "SAINT"
                        if not steFlag and word == "SAINTE":
                            steFlag = saint == "SAINTE"
                    elif len(saint) <= 3 and saint in word:
                        if not stFlag and word == "ST":
                            stFlag = saint == "ST"
                        if not steFlag and word == "STE":
                            steFlag = saint == "STE"

                if stFlag:
                    newStreet += "ST. "
                elif steFlag:
                    newStreet += "STE. "
                else:
                    newStreet += word + "-"

        if "||" in address.street:  # account for the case that a street was mistaken as a suffix

            if address.street[:address.street.index("|")] == address.street[address.street.index("||") + 2:]:
                newStreet = address.street[:address.street.index("|")]
            else:
                newStreet = address.street[address.street.index("||") + 2:]

    if not address.street.isupper():
        newStreet = newStreet.upper()
        address.flag.addStrFlag("FORMAT")

    if len(newStreet) != 0:
        newStreet = newStreet.replace("HIGHWAY", "HWY").replace("HIWAY", "HWY")
        temp = newStreet.split()
        for word in temp:
            if word == "RTE":
                temp.insert(temp.index("RTE"), "ROUTE")
                temp.remove("RTE")
        newStreet = " ".join(temp)
        address.street = newStreet.strip("-")
    else:
        temp = address.street.split()
        for word in temp:
            if word == "RTE":
                temp.insert(temp.index("RTE"), "ROUTE")
                temp.remove("RTE")
        address.street = " ".join(temp)
        address.street = address.street.replace("HIGHWAY", "HWY").replace("HIWAY", "HWY").strip("-")

    # SUFFIX #############################################

    if len(address.suffix) == 0 and "||" not in address.street and not isPO:
        address.flag.addSufFlag("UNDEFINED")

    if not address.suffix.isalpha() and len(address.suffix) != 0:
        address.suffix = removeSymbols(address.suffix)
        address.flag.addSufFlag("SYM")

    # based on the suffix, determine if it should be short or not

    if address.altSuffix in shortStreets:
        if address.suffix != shortStreets[address.altSuffix] and (not (
                address.street[-1].isalpha() and address.street[:-1].isdigit() and len(
            address.street.split()) == 1) or checkOrdinal(address.street)):
            newSuffix += shortStreets[address.altSuffix]
            address.flag.addSufFlag(address.altSuffix)
        elif address.street[-1].isalpha() and address.street[:-1].isdigit() and len(
                address.street.split()) == 1 and not checkOrdinal(address.street):
            newSuffix += address.altSuffix
    else:
        if len(address.suffix) != len(address.altSuffix):
            newSuffix += address.altSuffix
            address.flag.addSufFlag(address.altSuffix)
        else:
            newSuffix += address.suffix

    if address.french and len(address.street) != 0 and len(address.altSuffix) != 0 and len(address.number) != 0:
        numIndex = -1
        for i in range(len(address.original)):
            if len(address.original[i]) > 0:
                if address.original[i][-1] == address.number[-1]:
                    numIndex = i
        if numIndex != -1 and numIndex < len(address.original) - 1:
            if removeSymbols(address.original[numIndex + 1]) != address.suffix:
                address.flag.addSufFlag("STRUCT")

    if len(newSuffix) != 0:
        address.suffix = newSuffix

    # EXTERNAL INFO ######################################
    if not address.extra.isnumeric() and (len(address.extra) > 1 or len(address.extra.split()) > 1) and (
            isPO and address.external):
        address.flag.addAddrFlag("EXCESS")

    if isPO and len(address.street) == 0 == len(address.number) == len(address.suffix) == len(address.direction):
        address.street = " ".join([address.extra, address.number])
        address.number = ""
        address.extra = ""

    if len(address.extra) > 40: # attempt to trim ext. info
        address.extra = trimExtInfo(address.extra)


def trimExtInfo(extra):
    """
    Attempt to shorten external info to meet character limits
    :param extra: External info string
    :return: The external info string
    """
    new = ""
    for word in extra:
        if word == "FLOOR":
            new += "FL"
        elif word == "APARTMENT":
            new += "APT"
        elif checkOrdinal(word):
            if "IEME" in word or "IE" in word or "IER" in word:
                new += word[:word.find("I")] + "E"
            elif "EME" in word:
                new += word[:word.find("E")] + "E"
        elif word == "BUREAU":
            new += "BUR"
        else:
            new += word
    new = new.replace("BUILDING", "BLDG").replace("APARTMENT", "APT").replace("NIVEAU", "NIV").replace("STATION", "STN").replace("PARK", "PK")
    return new


def save_rules():
    """
    Save the rules onto a text file
    """
    with open("Rules.txt", "w") as saveFile:
        json.dump({"SUFFIXES":suffixes, "SUFF_PREF":shortStreets, "EXT":extras}, saveFile, indent=3, sort_keys=True,
                  ensure_ascii=False)

    print("Rules Saved")


def load_rules(suffixes, extras, shortStreets):
    """
    Load the rules from txt file
    :param suffixes: Suffix rules
    :param extras: External info rules
    :param shortStreets: Suffixes with alternate versions
    """
    with open("Rules.txt") as jsonFile:
        rules = json.load(jsonFile)

    suffixes.update(rules["SUFFIXES"])
    extras.extend(rules["EXT"])
    shortStreets.update(rules["SUFF_PREF"])


def write(fileName, addresses):
    """
    Writes new addresses and its flag to a new excel spreadsheet
    :param fileName: File Name to be used
    :param addresses: List of addresses
    """
    writeWb = Workbook(write_only=True)
    sheet = writeWb.create_sheet("Flags", 0)
    while os.path.isfile(fileName):
        fileName = fileName[:fileName.find(".xlsx")] + " - Copy" + ".xlsx"
    sheet.append(["AddressLine1", "AddressLine2", "Flags for Program"])
    for address in addresses:
        if "INVALID" in address.flag.address or (
                "UNDEFINED" in address.flag.number and "UNDEFINED" in address.flag.street and "UNDEFINED" in address.flag.suffix):
            sheet.append([" ".join(address.original), address.extra, "Invalid, Thus Unchanged"])
        else:
            sheet.append([str(address), address.extra])

    writeWb.save(filename=fileName)

    writeWb.close()
    gui.popup(str(len(addresses)) + " cleaned addresses written to " + os.path.abspath(fileName) + '. Took ' + str(
        timer() - start) + " seconds to finish.")
    print("FINISHED")


def debugWrite(addresses):
    writeWb = Workbook(write_only=True)
    sheet = writeWb.create_sheet("Flags", 0)

    fileName = "testclean.xlsx"
    sheet.append(["AddressLine1", "AddressLine2", "Flags for Program"])
    for address in addresses:
        if "INVALID" in address.flag.address or (
                "UNDEFINED" in address.flag.number and "UNDEFINED" in address.flag.street and "UNDEFINED" in address.flag.suffix):
            sheet.append([" ".join(address.original), address.extra, "Invalid, Thus Unchanged"])
        else:
            sheet.append([str(address), address.extra, str(address.flag), " ".join(address.original)])

    writeWb.save(filename=fileName)

    writeWb.close()
    gui.popup(str(len(addresses)) + " cleaned addresses written to " + os.path.abspath(fileName) + '. Took ' + str(
        timer() - start) + " seconds to finish.")
    print("FINISHED")


def createTemplate():
    """
    Generates a template for address reading for users
    """
    writeWb = Workbook()
    header = NamedStyle(name="Title")
    note = NamedStyle(name="Note")
    exception = NamedStyle(name="Exception")

    info = writeWb.active  # first sheet
    info.title = "About"
    template = writeWb.create_sheet("Template", 1)
    templateBlank = writeWb.create_sheet("Blank Template", 2)

    header.font = Font(bold=True, size=11)
    header.fill = PatternFill(fill_type='solid', start_color="00FFFF00", end_color="00FFFF00")
    template['A1'] = "Unique ID"
    template['B1'] = "AddressLine1"
    template['C1'] = "AddressLine2"
    template['D1'] = "Province"
    templateBlank['A1'] = "Unique ID"
    templateBlank['B1'] = "AddressLine1"
    templateBlank['C1'] = "AddressLine2"
    templateBlank['D1'] = "Province"

    for cell in template["1"]:
        if cell.value:
            cell.style = header

    for cell in templateBlank["1"]:
        if cell.value:
            cell.style = header

    note.font = Font(italic=True, size=11)
    template['B2'] = "Main Address (North America only)"
    template['C2'] = "External Information"
    template['D2'] = "Helps determine whether Address is FR or EN"

    for cell in template["2"]:
        if cell.value:
            cell.style = note

    template['B3'] = "123 SAMPLE STREET"
    template['C3'] = "PO BOX 23"
    template['D3'] = "ON"
    template['B4'] = "432 RUE MONTREAL"
    template['C4'] = "APT B"
    template['D4'] = "QC"
    template['B5'] = "CP 30"
    template['D5'] = "QC"

    exception.font = Font(italic=True, size=11, bold=True)
    template["B8"] = "Exceptions (must be fixed manually or left alone)"
    template["B9"] = "GPS Coordinates"
    template["B10"] = "Directions to location"

    for cell in template["8"]:
        if cell.value:
            cell.style = exception

    # set the row length to the longest value
    dimens = {}
    dimensB = {}
    for row in template.rows:
        for cell in row:
            if cell.value:
                dimens[cell.column_letter] = max((dimens.get(cell.column_letter, 0)), len(str(cell.value)))
    for row in templateBlank.rows:
        for cell in row:
            if cell.value:
                dimensB[cell.column_letter] = max((dimensB.get(cell.column_letter, 0)), len(str(cell.value)))

    for col, value in dimens.items():
        template.column_dimensions[col].width = value+5

    for col, value in dimensB.items():
        templateBlank.column_dimensions[col].width = value + 5

    # write the file and let the user know where to find the file
    writeWb.save(filename='AddressTemplate.xlsx')
    gui.popup("Template saved to " + os.path.abspath("AddressTemplate.xlsx"))
    writeWb.close()


# inner class for storing flags
class Flag:
    def __init__(self):
        self._number = set() # errors with number
        self._suffix = set() # errors with suffix
        self._street = set() # errors with street
        self._direction = set() # errors with direction
        self._address = set() # errors with overall address

    def addNumFlag(self, flag):
        self.number.add(flag)

    def addDirFlag(self, flag):
        self.direction.add(flag)

    def addStrFlag(self, flag):
        self.street.add(flag)

    def addSufFlag(self, flag):
        self.suffix.add(flag)

    def addAddrFlag(self, flag):
        self.address.add(flag)

    @property
    def number(self):
        return self._number

    @property
    def suffix(self):
        return self._suffix

    @property
    def street(self):
        return self._street

    @property
    def direction(self):
        return self._direction

    @property
    def address(self):
        return self._address

    @number.setter
    def number(self, number):
        pass

    @suffix.setter
    def suffix(self, suffix):
        pass

    @street.setter
    def street(self, street):
        pass

    @direction.setter
    def direction(self, direct):
        pass

    @address.setter
    def address(self, address):
        pass

    def isValid(self):
        return len(self._number) == len(self._direction) == len(self._suffix) == len(self._street) == len(
            self._address) == 0

    def __str__(self):
        if self.isValid():
            return "VALID"
        else:
            flags = [[], [], [], [], []]
            for flag in self._number:
                flags[0].append(flag)
            for flag in self._street:
                flags[1].append(flag)
            for flag in self._suffix:
                flags[2].append(flag)
            for flag in self._direction:
                flags[3].append(flag)
            for flag in self._address:
                flags[4].append(flag)
            message = ""
            i = 0
            for word in ["N", "ST", "SF", "D", "A"]:
                message += word + "["
                for flag in flags[i]:
                    message += flag + "/"
                i += 1
                message.rstrip("/")
                message += "]"
            return message


# inner class for addresses
class Address:
    def __init__(self, original):
        self._number = "" # number of address
        self._original = original # original address before any simplifying
        self._street = "" # street name
        self._suffix = "" # street suffix
        self._altSuffix = "" # full version of suffix if the suffix itself is short
        self._direction = "" # cardinal direction
        self._french = False # french address
        self._flag = Flag() # any errors with the address
        self._extra = "" # external info
        self._external = False  # if the external info is an address
        self._ordinal = False  # if ordinal numbers exist in the street
        self._po = False # if the address is a po box
        self._suffixNumber = "" # usually with rural roads, can have number attached with the suffix

    # GETTERS AND SETTERS
    @property
    def number(self):
        return self._number

    @number.setter
    def number(self, number):
        self._number = number

    @property
    def street(self):
        return self._street

    @street.setter
    def street(self, street):
        self._street = street

    @property
    def suffix(self):
        return self._suffix

    @suffix.setter
    def suffix(self, suffix):
        self._suffix = suffix

    @property
    def direction(self):
        return self._direction

    @direction.setter
    def direction(self, direction):
        self._direction = direction

    @property
    def flag(self):
        return self._flag

    @flag.setter
    def flag(self, flag):
        self._flag = flag

    def isValid(self):
        return self._flag.isValid()

    @property
    def original(self):
        return self._original

    @original.setter
    def original(self, original):
        self._original = original

    @property
    def french(self):
        return self._french

    @french.setter
    def french(self, value):
        self._french = value

    @property
    def external(self):
        return self._external

    @external.setter
    def external(self, external):
        self._external = external

    @property
    def altSuffix(self):
        return self._altSuffix

    @altSuffix.setter
    def altSuffix(self, altSuffix):
        self._altSuffix = altSuffix

    @property
    def extra(self):
        return self._extra

    @extra.setter
    def extra(self, extra):
        self._extra = extra

    @property
    def ordinal(self):
        return self._ordinal

    @ordinal.setter
    def ordinal(self, ordinal):
        self._ordinal = ordinal

    @property
    def po(self):
        return self._po

    @po.setter
    def po(self, po):
        self._po = po

    @property
    def suffixNumber(self):
        return self._suffixNumber

    @suffixNumber.setter
    def suffixNumber(self, num):
        self._suffixNumber = num

    def __str__(self):
        if self._french and not self._ordinal:  # add ordinal later
            return " ".join(" ".join([self._number, self._suffix.strip(), self._street.strip().rstrip("-"),
                                      self._suffixNumber, self._direction]).split()).strip().rstrip(",")
        return " ".join(" ".join(
            [self._number, self._street.strip().rstrip("-"), self._suffix, self._direction,
             self._suffixNumber]).split()).strip().rstrip(
            ",")


# ******************************************************* MAIN ********************************************************

print("Initializing program")

extras = []
suffixes = {}
shortStreets = {}
streets = []
userStreets = []
load_rules(suffixes, extras, shortStreets)

for street in suffixes:
    streets.append(street)

# display the streets
for i in range(len(streets)):
    if streets[i] in shortStreets:
        streets[i] += "->" + shortStreets[streets[i]]


# GUI INITIALIZATION ############################################################################
gui.theme("DarkTeal12")
# gui layouts
aboutLayout = [[gui.Text("Welcome to the Address Cleaner!", font="Arial 15 bold")],
               [gui.Text("\nDeveloped by Albert Quon (aquon095@uottawa.ca)")],
               [gui.Text("Last Updated: 2020-08-25\n")],
               [gui.Text("Click 'Create Template' to generate a template of the required format")],
               [gui.Button("Create Template")]]
fileLayout = [[gui.Text("NOTE: Reading a file can take up to 5 minutes depending on the size, \nthis window may not respond during scanning and writing")],
              [gui.Text("The sheet with the inputted name must be in the format given in the template")],
              [gui.Text("Enter the sheet name:"), gui.InputText(key='SHEET_IN'), gui.Button("Submit Sheet")],
              [gui.Text("Enter file name (must be in the same folder as program)"),
               gui.InputText(key="FILE_IN", default_text=".xlsx")],
              [gui.Button("Read File")], [gui.Button("Exit")]]

menuRule = ['File', ['Remove Rule']]
suffixLayout = [[gui.Text("Suffix (AddressLine1) Rules", font="Arial 13 bold")],
                [gui.Text("View, add, or delete suffix rules. NOTE: New rules may result in invalid addresses due to lack of testing.")],
                [gui.Text("NOTE: Deleting any program default suffix rules will result in unreliable results.")],
                [gui.Text("Suffix rules are used to standardize words related to suffixes in streets to be consistent.")],
                [gui.Text("\nEnter a Suffix to add")], [gui.InputText(key="SUFF_IN")],
                [gui.Text("Enter preferred version of the street if needed")],
                [gui.InputText(key="ALT_SUFF_IN")], [gui.Button("Submit Suffix Rule")],
                [gui.Text("\nDEFAULT suffixes in addresses", auto_size_text=True)],
                [gui.Listbox(streets, size=(25, 10), key="SUFFIXES", right_click_menu=menuRule, enable_events=True)],
                [gui.Text("\nRules added this session")], [
                    gui.Listbox(userStreets, size=(20, 10), key="USER_SUFF", right_click_menu=menuRule,
                                enable_events=True)],
                [gui.Button("Save Rules")]]

extLayout = [[gui.Text("External Infomation (AddressLine2) Rules", font="Arial 13 bold")],
             [gui.Text("View, add, or delete external rules. NOTE: New rules may result in invalid addresses due to lack of testing.")],
             [gui.Text("External rules are used to recognize words that are to be separated from the main address to external information.")],
             [gui.Text("\nEnter a word to be recognized as external information")], [gui.InputText(key="EXT_IN")], [gui.Button("Submit External Rule")],
             [gui.Text("\nHere are the current list of external info in addresses", auto_size_text=True)],
             [gui.Listbox(extras, size=(25, 10), key="EXTERNALS", right_click_menu=menuRule, enable_events=True)],
             [gui.Button("Save")]]

debugLayout = [[gui.Button("Debug (no writing)")], [gui.Button("Debug Write")], [gui.Button("Debug Load")]]

layout = [[gui.TabGroup(
    [[gui.Tab('About', aboutLayout)], [gui.Tab('File', fileLayout)], [gui.Tab('Suffix', suffixLayout)],
     [gui.Tab('External Info', extLayout)], [gui.Tab('Debug', debugLayout, visible=False)]])]]
window = gui.Window("Address Cleaner", layout, resizable=True)
fileName = ""
# GUI LOOP ############################################################################
try:
    terminate = False
    sheetName = "" # default name
    while not terminate: # handle all events
        event, values = window.read()
        if event == gui.WIN_CLOSED or event == "Exit":
            save_rules()
            terminate = True

        if event == "Debug (no writing)" or event == "Debug Write":
            fileName = "debug.xlsx"
            sheetName = "Sheet1"  # default name
            #terminate = True
            start = timer()
            workbook = load_workbook(filename=fileName, read_only=True)  # 110 sec load time with read_only false
            sheetTest = workbook[sheetName]

            print("Workbook finished loading")
            print(timer() - start)
            if event == "Debug Write":
                debugWrite(scan(sheetTest))
            elif event == "Debug (no writing)":
                scan(sheetTest)

            workbook.close()
        if event == "Debug Load":
            load_rules(suffixes, extras, shortStreets)

        if event == "Read File":
            start = timer()
            fileName = values["FILE_IN"]
            if os.path.isfile(fileName):
                validInput = True
            elif len(sheetName) == 0:
                validInput = False
                window.FindElement("SHEET_IN").Update(background_color="red")
                gui.popup("Please input a sheet name!")
            else:
                validInput = False
                window.FindElement("FILE_IN").Update(background_color="red")
                gui.popup("Invalid File Name!")
            if validInput:
                workbook = load_workbook(filename=fileName, read_only=True)  # 110 sec load time with read_only false
                sheet = workbook[sheetName]
                outFileName = fileName[:fileName.find(".xlsx")] + "Cleaned.xlsx"
                write(outFileName, scan(sheet))
                print("Workbook finished loading")
                print(timer() - start)
                #main()
                workbook.close()

        if event == "Create Template":
            createTemplate()

        if event == "Submit Suffix Rule":
            if len(values["ALT_SUFF_IN"]) > 0: # check if a preferred representation of a suffix has been added
                if len(values["SUFF_IN"]) > 0 and values["SUFF_IN"].isalpha():
                    if values["ALT_SUFF_IN"].isalpha() and\
                            (values["SUFF_IN"].upper() + "->" + values["ALT_SUFF_IN"].upper() not in streets and
                             values["SUFF_IN"].upper() + "->" + values["ALT_SUFF_IN"].upper() not in userStreets)\
                            and len(values["SUFF_IN"]) > len(values["ALT_SUFF_IN"]): # check if alt suffix is valid
                        shortStreets[values["SUFF_IN"].upper()] = values["ALT_SUFF_IN"].upper()
                        userStreets.append(values["SUFF_IN"].upper() + "->" + values["ALT_SUFF_IN"].upper())
                        factors = []
                        j=0
                        for i in range(len(values["SUFF_IN"])):
                            if j < len(values["ALT_SUFF_IN"]):
                                if values["SUFF_IN"][i].upper() == values["ALT_SUFF_IN"][j].upper():
                                    factors.append(1/len(values["ALT_SUFF_IN"])-1/len(values["ALT_SUFF_IN"])%0.01)
                                    j+=1
                                else:
                                    factors.append(0.07)
                            else:
                                factors.append(0.07)
                        suffixes[values["SUFF_IN"].upper()] = factors
                        values["ALT_SUFF_IN"] = ""
                    else:
                        if not values["ALT_SUFF_IN"].isalpha() and len(values["ALT_SUFF_IN"]) > 0:
                            gui.popup("Not a valid preferred suffix! No rules have been added.")
                    values["SUFF_IN"] = ""
                    values["ALT_SUFF_IN"] = ""
                elif not values["SUFF_IN"].isalpha():
                    gui.popup("Not a valid suffix!")
            else: # only a suffix has been given
                if len(values["SUFF_IN"]) > 0 and values["SUFF_IN"].isalpha() and values[
                    "SUFF_IN"].upper() not in streets and\
                        values["SUFF_IN"].upper() not in userStreets:
                    userStreets.append(values["SUFF_IN"].upper())
                    factors = []
                    for i in range(len(values["SUFF_IN"])):
                        factors.append(1/len(values["SUFF_IN"])-1/len(values["SUFF_IN"])%0.01)
                    factors[-1] += 0.01
                    suffixes[values["SUFF_IN"].upper()] = factors
                    values["SUFF_IN"] = ""
                elif not values["SUFF_IN"].isalpha():
                    gui.popup("Not a valid suffix!")
            streets.sort()
            window.FindElement('USER_SUFF').Update(values=userStreets)

        if event == "Remove Rule":
            # remove a rule based on what was inputted
            if len(values["SUFFIXES"]) > 0:
                if "->" in values["SUFFIXES"][0]:
                    suff = values["SUFFIXES"][0][:values["SUFFIXES"][0].find("->")]
                else:
                    suff = values["SUFFIXES"][0]

                suffixes.pop(suff)
                streets.remove(values["SUFFIXES"][0])
                window.FindElement('SUFFIXES').Update(values=streets)

                if suff in shortStreets:
                    shortStreets.pop(suff)
            elif len(values["USER_SUFF"]) > 0:
                if "->" in values["USER_SUFF"][0]:
                    suff = values["USER_SUFF"][0][:values["USER_SUFF"][0].find("->")]
                else:
                    suff = values["USER_SUFF"][0]

                userStreets.remove(values["USER_SUFF"][0])
                window.FindElement('USER_SUFF').Update(values=userStreets)

                if suff in shortStreets:
                    shortStreets.pop(suff)
            if len(values["EXTERNALS"]) > 0:
                extras.remove(values["EXTERNALS"][0])
                window.FindElement("EXTERNALS").Update(values=extras)

        if event == "Submit External Rule":
            if len(values["EXT_IN"]) > 0 and values["EXT_IN"].upper() and values["EXT_IN"].isalpha() and \
                    values["EXT_IN"] not in extras:

                extras.append(values["EXT_IN"].upper())
                extras.sort()
                window.FindElement("EXTERNALS").Update(values=extras)
                window.FindElement("EXT_IN").Update(value='')
            else:
                gui.popup("Not a proper rule!")

        if event == "Submit Sheet": #sheet name
            sheetName = values["SHEET_IN"]
            if len(sheetName) > 0:
                window.FindElement("SHEET_IN").Update(background_color="green")
            else:
                window.FindElement("SHEET_IN").Update(background_color="red")

        if event == "Save" or event == "Save Rules":
            save_rules()

    window.close()
except KeyError:
    gui.popup("Worksheet entered does not exist")
except PermissionError:
    gui.popup("Please close excel files associated with this program!")
except: # while not ideal, can be used to tell when something wrong happens
    traceback.print_exc()
    gui.popup("Error Occured")


# KNOWN BUGS ###############################################################
# gps coordinates and direction instructions
# words similar to ext rules are put into ext info instead of street

# PLANS ####################################################################
# recognize things the programs cannot cleanup and leave it alone
# - notably: gps coordinates and direction instructions