from openpyxl import load_workbook, Workbook
#from openpyxl.cell import WriteOnlyCell
#import openpyxl.cell
from timeit import default_timer as timer
import os.path
from datetime import datetime


def initializeAssets():
    """
    Initialize the assets from the report
    :param sheet: The worksheet of the report
    :param columns: Columns to scan through
    :return: List of all assets
    """
    assets = []
    sheet = workbook["Combined Device List"]
    assetCol = findCol("Found Device Name\n(workstations)", sheet["5"])
    userCol = findCol("Owner Name Full\n(from AD)", sheet["5"])
    statusCol = findCol("Classification", sheet["5"])
    typeCol = findCol("Type", sheet["5"])
    dateCol = findCol("Last Logon Date", sheet["5"])


    for row in sheet.iter_rows(min_row=6):
        name = str(row[assetCol].value).strip("_").upper()
        if (name.isdigit() and len(name) == 5) or (("ASSET-" in name) and (
        ("Workstation" == str(row[typeCol].value) or "Workstation (Apple)" == str(row[typeCol].value))) and len(
                name[6:]) == 5 and name[6:].isalnum()) or\
                "Workstation (Apple)" == str(row[typeCol].value):
            if "ASSET-" in name:
                name = name[6:]
        else:
            name = ""

        if len(name) > 0:
            asset = Asset(name)
            if userCol != -1:
                user = " ".join(str(row[userCol].value).strip().split()).upper()
            else:
                user = "N/A"
            if len(user) == 0 or user is None or "NONE" in user or "N/A" in user or "NOT IN AD" in user:
                asset.user["Report"] = "N/A"
            else:
                asset.user["Report"] = user

            asset.flags["STATUS"] = ""
            asset.flags["USER"] = ""
            status = str(row[statusCol].value).lower()
            asset.active["Report"] = "active" in status and "?" not in status and "inactive" not in status

            asset.status["Report"] += status
            asset.appleUnverified = "tbd, check with carl for jamf info" in status
            asset.loginDate = str(row[dateCol].value).lower()
            if "2020" not in asset.loginDate and asset.loginDate is not None and asset.loginDate != "none":
                asset.specialNote += "[Last login date was long ago: " + asset.loginDate + "]"

            if "inactive" in asset.status["Report"] and "N/A" not in asset.user["Report"] and not "tbd" in status:
                asset.flags["USER"] += "[User with inactive asset]"
            asset.report = True

            asset.verified = "same" in str(row[userCol + 4].value).strip().lower()

            if asset.verified:
                asset.specialNote += "[Verified user]"

            if asset not in assets:  # or "decom" not in str(row[userCol+2].value).lower():
                assets.append(asset)
            else:
                print("ASSET CONFLICT")

    return assets


def findCol(column, sheet):
    """
    Finds the column in a spreadsheet with a given name
    :param column: Name of the column
    :param sheet: Worksheet to go through
    :return: index of the column
    """
    for cell in sheet:
        if cell.value == column:
            col = sheet.index(cell)
            return col
    return -1


def compareAndMerge(sheet, columns, assets, isEmerge):
    """
    # one loop, go through all the necessary columns, then flag if not consistent
    :param sheet: the spreadsheet
    """

    assetCol = findCol(columns[0], sheet["1"])
    userCol = findCol(columns[1], sheet["1"])
    statusCol = findCol(columns[2], sheet["1"])

    # DATA VALIDATION - check if the address is valid itself
    # if the address is valid, check if it is consistent
    for row in sheet.iter_rows(min_row=2):
        name = str(row[assetCol].value).strip()
        if (name.isdigit() and len(name) == 5 and int(name) >= 80000) or (
                ("ASSET-" in name) and len(name[6:]) == 5 and name[6:].isalnum()):
            if "ASSET-" in name:
                name = name[6:]
        else:
            name = ""

        if len(name) > 0:
            asset = Asset(name)
            if asset in assets:
                asset = assets[assets.index(asset)]
                if not isEmerge:
                    asset.jde = True
                    asset.active["JDE"] = True
                else:
                    asset.emerge = True

                if userCol != -1:
                    user = " ".join(str(row[userCol].value).strip().split()).upper()
                    if len(user) == 0 or user is None or user == "NONE" or "N/A" in user:
                        user = "N/A"

                    if isEmerge:
                        asset.user["Emerge"] = user
                    else:
                        asset.user["JDE"] = user

                if statusCol != -1:
                    status = str(row[statusCol].value).lower()

                    if isEmerge:
                        asset.active["Emerge"] = "active" in status
                        asset.status["Emerge"] = status
                        #print(asset.active["Emerge"])
                    else:
                        asset.active["JDE"] = "active" in status
                        asset.status["JDE"] = status

                    # if ("disposed" in status or "returned" in status) and (
                    #         asset.active["JDE"] or asset.active["Emerge"] or asset.active["Report"]):
                    #     asset.specialNote += status
                    #     if isEmerge:
                    #         asset.specialNote += "(E)"
                    #     else:
                    #         asset.specialNote += "(J)"

            else:

                if not isEmerge:
                    asset.jde = asset.active["JDE"] = True
                else:
                    asset.emerge = True

                if userCol != -1:
                    user = " ".join(str(row[userCol].value).strip().split()).upper()
                    if len(user) == 0 or user is None or user == "NONE" or "N/A" in user:
                        user = "N/A"
                    if isEmerge:
                        asset.user["Emerge"] = user
                    else:
                        asset.user["JDE"] = user

                if statusCol != -1:
                    status = str(row[statusCol].value).strip().lower()
                    if isEmerge:
                        asset.active["Emerge"] = "active" in status
                        asset.status["Emerge"] = status

                        #print(asset.active["Emerge"])
                    else:
                        asset.active["JDE"] = "active" in status
                        asset.status["JDE"] = status
                    #if not isEmerge: # and "disposed" not in str(row[statusCol].value).strip():
                    # if any(msg in status for msg in ["disposed", "unverified", "pre-install", "cage"]):
                    #     asset.specialNote += status
                    #     if isEmerge:
                    #         asset.specialNote += "(E)"
                    #     else:
                    #         asset.specialNote += "(J)"

                asset.flags["USER"] = ""
                asset.flags["STATUS"] = ""

                if statusCol != -1 and userCol != -1 and isEmerge:
                    #print(asset.user["Emerge"], asset.status["Emerge"])
                    #print(asset.status["Emerge"], any(msg in asset.status["Emerge"] for msg in ["disposed", "unverified", "pre-install", "cage"]))
                    if any(msg in asset.status["Emerge"] for msg in ["disposed", "unverified", "pre-install", "cage"]):

                        if any("N/A" not in asset.user[sheet] and len(asset.user[sheet]) > 0 for sheet in asset.user):
                            asset.flags["USER"] += "[User with inactive asset]"
                        else:
                            asset.valid = True

                #if not isEmerge: # if in JDE, make sure it is not a monitor
                descA = str(row[3].value).lower()
                descB = str(row[4].value).lower()

                if any(word in descA for word in ["monitor", "lcd", '"', "scanner", "led", "phone"]) or any(word in descB for word in ["monitor", "lcd", '"', "scanner", "led", "phone"]):
                    asset.workstation = False
                    asset.specialNote += "[Not a workstation]"

                assets.append(asset)


def flagIssues(assets):
    users = set()
    duplicateCount = differentUser = differentStatus = 0
    for asset in assets:
        userPresent = set()
        assetActives = set()

        if asset.jde:
            assetActives.add(asset.active["JDE"])
            userPresent.add(asset.user["JDE"])
        if asset.emerge:
            assetActives.add(asset.active["Emerge"])
            userPresent.add(asset.user["Emerge"])
        if asset.report:
            assetActives.add(asset.active["Report"])
            userPresent.add(asset.user["Report"])

        if len(assetActives) > 1:
            if "inactive" in asset.status["Report"] and ((asset.jde and asset.active["JDE"]) or (asset.report and asset.active["Report"])):
                asset.flags["STATUS"] += "[JDE/Emerge may need update]"
            else:
                asset.flags["STATUS"] += "[Inconsistent Status]"
            differentStatus += 1

        if (len(asset.loginDate) > 0 and asset.loginDate != "none" and asset.loginDate is not None) and ("2020" not in asset.loginDate and "2019" not in asset.loginDate) and not asset.isValidInactive():
            print(asset.loginDate)
            asset.flags["STATUS"] += "[Asset should be inactive]"

        if len(userPresent) > 1 and not asset.verified:
            asset.flags["USER"] += "[Inconsistent User]"
            differentUser += 1

        assumedLength = len(users)
        assetUsers = set()
        for sheet in asset.user:
            if not (asset.user[sheet] == "N/A" or len(asset.user[sheet]) == 0):
                assetUsers.add(asset.user[sheet])
        assumedLength += len(assetUsers)

        if asset.isValidInactive() and len(assetUsers) > 1:
            asset.flags["USER"] += "[User with inactive asset]"

        users = users.union(assetUsers)
        if len(users) != assumedLength and len(assetUsers) != 0:
            asset.flags["USER"] += "[User with multiple devices]"
            duplicateCount += 1
    print(duplicateCount, "Users with multiple devices")
    print(differentUser, "Devices with multiple users")
    print(differentStatus, "Devices with multiple statuses")


def countAssetCategories(assets):
    numEmerge = numEmergeJDE = numJDE = numJDEReport = numReport = numReportEmerge = numAll = 0
    for asset in assets:
        if asset.emerge:
            numEmerge += 1
        if asset.report:
            numReport += 1
        if asset.jde:
            numJDE += 1
        if asset.jde and asset.emerge:
            numEmergeJDE += 1
        if asset.jde and asset.report:
            numJDEReport += 1
        if asset.report and asset.emerge:
            numReportEmerge += 1
        if asset.report and asset.emerge and asset.jde:
            numAll += 1

    return (numEmerge, numJDE, numReport, numJDEReport, numReportEmerge, numEmergeJDE, numAll)


def write(assets):
    """
    Creates a new excel spreadsheet and writes out the analysis
    :param assets: The assets that were analyzed
    """
    writeWb = Workbook(write_only=True)

    rawSheet = writeWb.create_sheet("Raw Data", 0)
    activeSheet = writeWb.create_sheet("Active Assets", 1)
    inactiveSheet = writeWb.create_sheet("Inactive", 2)
    inconsistSheet = writeWb.create_sheet("Inconsistent", 3)
    appleSheet = writeWb.create_sheet("Unverifed Apple Assets", 4)
    unknownSheet = writeWb.create_sheet("Unverified Assets", 5)
    nonCompSheet = writeWb.create_sheet("Non-workstations", 6)

    fileName = "AssetAnalysis.xlsx"
    # Headers
    rawSheet.row_dimensions["1"].width = 25 * 4.0
    activeSheet.row_dimensions["1"].width = 25 * 4.0
    inactiveSheet.row_dimensions["1"].width = 25 * 4.0
    inconsistSheet.row_dimensions["1"].width = 25 * 4.0

    rawSheet.append(
        ["Asset", "JDE User", "In JDE", "Active in JDE", "JDE Status", "Emerge User", "In Emerge", "Active in Emerge",
         "Emerge Status",
         "Report User", "In Live Data", "Active in Live Source", "Live Source Status", "USER Flags", "STATUS Flags",
         "Notes"])
    activeSheet.append(
        ["Asset", "JDE User", "In JDE", "Active in JDE", "JDE Status", "Emerge User", "In Emerge", "Active in Emerge",
         "Emerge Status",
         "Report User", "In Live Data", "Active in Live Source", "Live Source Status", "USER Flags", "STATUS Flags",
         "Notes"])
    inactiveSheet.append(
        ["Asset", "JDE User", "In JDE", "Active in JDE", "JDE Status", "Emerge User", "In Emerge", "Active in Emerge",
         "Emerge Status",
         "Report User", "In Live Data", "Active in Live Source", "Live Source Status", "USER Flags", "STATUS Flags",
         "Notes"])
    inconsistSheet.append(
        ["Asset", "JDE User", "In JDE", "Active in JDE", "JDE Status", "Emerge User", "In Emerge", "Active in Emerge",
         "Emerge Status",
         "Report User", "In Live Data", "Active in Live Source", "Live Source Status", "USER Flags", "STATUS Flags",
         "Notes"])
    appleSheet.append(
        ["Asset", "JDE User", "In JDE", "Active in JDE", "JDE Status", "Emerge User", "In Emerge", "Active in Emerge",
         "Emerge Status",
         "Report User", "In Live Data", "Active in Live Source", "Live Source Status", "USER Flags", "STATUS Flags",
         "Notes"])
    unknownSheet.append(["Asset", "JDE User", "In JDE", "Active in JDE", "JDE Status", "Emerge User", "In Emerge", "Active in Emerge",
         "Emerge Status",
         "Report User", "In Live Data", "Active in Live Source", "Live Source Status", "USER Flags", "STATUS Flags",
         "Notes"])
    nonCompSheet.append(["Asset", "JDE User", "In JDE", "Active in JDE", "JDE Status", "Emerge User", "In Emerge", "Active in Emerge",
         "Emerge Status",
         "Report User", "In Live Data", "Active in Live Source", "Live Source Status", "USER Flags", "STATUS Flags",
         "Notes"])

    for asset in assets:
        # if len(asset.flags["USER"]) == 0:
        #     asset.flags.pop("USER")
        # if len(asset.flags["STATUS"]) == 0:
        #     asset.flags.pop("STATUS")

        if asset.jde:
            jde = "In JDE"
        else:
            jde = "NOT in JDE"

        if asset.report:
            report = "In Live Source"
        else:
            report = "NOT in Live Source"

        if asset.emerge:
            emerge = "In Emerge"
        else:
            emerge = "NOT in Emerge"

        rawSheet.append(
            [asset.name, asset.user["JDE"], jde, asset.active["JDE"], asset.status["JDE"], asset.user["Emerge"], emerge,
             asset.active["Emerge"], asset.status["Emerge"], asset.user["Report"], report, asset.active["Report"],
             asset.status["Report"], asset.flags["USER"], asset.flags["STATUS"], asset.specialNote])
        if asset.appleUnverified and asset.report:
            appleSheet.append(
                [asset.name, asset.user["JDE"], jde, asset.active["JDE"], asset.status["JDE"], asset.user["Emerge"],
                 emerge,
                 asset.active["Emerge"], asset.status["Emerge"], asset.user["Report"], report, asset.active["Report"],
                 asset.status["Report"], asset.flags["USER"], asset.flags["STATUS"], asset.specialNote])
        elif not asset.workstation:
            nonCompSheet.append(
                [asset.name, asset.user["JDE"], jde, asset.active["JDE"], asset.status["JDE"], asset.user["Emerge"],
                 emerge,
                 asset.active["Emerge"], asset.status["Emerge"], asset.user["Report"], report, asset.active["Report"],
                 asset.status["Report"], asset.flags["USER"], asset.flags["STATUS"], asset.specialNote])
        elif "tbd" in asset.status["Report"]:
            unknownSheet.append([asset.name, asset.user["JDE"], jde, asset.active["JDE"], asset.status["JDE"], asset.user["Emerge"],
                 emerge,
                 asset.active["Emerge"], asset.status["Emerge"], asset.user["Report"], report, asset.active["Report"],
                 asset.status["Report"], asset.flags["USER"], asset.flags["STATUS"], asset.specialNote])
        elif asset.isValidActive():
            activeSheet.append(
                [asset.name, asset.user["JDE"], jde, asset.active["JDE"], asset.status["JDE"], asset.user["Emerge"],
                 emerge,
                 asset.active["Emerge"], asset.status["Emerge"], asset.user["Report"], report, asset.active["Report"],
                 asset.status["Report"], asset.flags["USER"], asset.flags["STATUS"], asset.specialNote])
        elif asset.isValidInactive():
            inactiveSheet.append(
                [asset.name, asset.user["JDE"], jde, asset.active["JDE"], asset.status["JDE"], asset.user["Emerge"],
                 emerge,
                 asset.active["Emerge"], asset.status["Emerge"], asset.user["Report"], report, asset.active["Report"],
                 asset.status["Report"], asset.flags["USER"], asset.flags["STATUS"], asset.specialNote])
        else:
            inconsistSheet.append(
                [asset.name, asset.user["JDE"], jde, asset.active["JDE"], asset.status["JDE"], asset.user["Emerge"],
                 emerge,
                 asset.active["Emerge"], asset.status["Emerge"], asset.user["Report"], report, asset.active["Report"],
                 asset.status["Report"], asset.flags["USER"], asset.flags["STATUS"], asset.specialNote])

    writeWb.save(filename=fileName)
    writeWb.close()


# inner class for an asset
class Asset:
    def __init__(self, name="", user=""):
        self.name = name
        self.user = {"Emerge":"", "JDE":"", "Report":""}
        self.emerge = False
        self.report = False
        self.jde = False
        self.active = {"Emerge": False, "JDE": False, "Report": False}
        self.status = {"Emerge":"", "JDE":"", "Report":""}
        self.flags = {}
        self.verified = False
        self.appleUnverified = False
        self.specialNote = ""
        self.workstation = True
        self.altUser = ""
        self.loginDate = ""

    @property
    def user(self):
        return self._user

    @user.setter
    def user(self, user):
        self._user = user



    def isValidActive(self):
        return (self.jde or self.emerge) and self.active["Report"] and self.report and self.verified

    def isValidInactive(self):
        return (self.report and not self.active["Report"] and "inactive" in self.status["Report"]) or (not self.report and not self.jde and self.emerge and not self.active["Emerge"])#and "2020" in self.loginDate

    def __str__(self):
        return self.name

    def __eq__(self, other):
        return other.name == self.name

    def __contains__(self, item):
        return any(item == self.name)

    def __hash__(self):
        return hash(self.name)

    def __repr__(self):
        return "Asset{}".format([self.name, self._user, self.active, self.emerge, self.jde, self.report])


print("Initializing program")
start = timer()
emergeFile = "redacted.xlsx"
jdeFile = "redacted.xlsx"
reportFile = "report.xlsx"

print("Opening workbook")
workbook = load_workbook(filename=reportFile, read_only=True)

print("Workbook finished loading")
print(timer() - start)
print("Scanning sheets")
assets = initializeAssets()
print(len(assets))
print("Workbook closed")
workbook.close()

print("Opening workbook")
workbook = load_workbook(filename=jdeFile, read_only=True)
print("Workbook finished loading")
print(timer() - start)
print("Scanning sheets")
compareAndMerge(workbook["Sheet1"], ("Name", "Device Owner Name", "Equipment Status"), assets, False)
print(len(assets))
print("Workbook closed")
workbook.close()

print("Opening workbook")
workbook = load_workbook(filename=emergeFile, read_only=True)
print("Workbook finished loading")
print(timer() - start)
print("Scanning sheets")
compareAndMerge(workbook["redacted"], ("Asset Tag", "Employee Name", "Asset Status"), assets, True)
print(len(assets))
print("Workbook closed")
workbook.close()

flagIssues(assets)
print("FINISHED WITH ASSETS")
print(timer() - start)

print(countAssetCategories(assets))

print(len(assets))
write(assets)
print("FINISHED AT ", end="")
now = datetime.now()
current_time = now.strftime("%H:%M")
print(current_time)

# Bucket 1 (xxxx)
# > Active all good EUDs
# > Active in live data source + present in JDE/Emerge + validated user
# > Possible misspelling, yet EUD-User mapping is ok (update JDE/Emerge with AD spelling)
#  
# Bucket 2 (yyyy)
# > Inconsistent EUDs
# > Active in live data source + present in JDE/Emerge + unverified user (reassigned/manager?)
# > Active in live data source + not-present in JDE/Emerge
#  
# Bucket 3 (zzzz)
# > InActive EUD
# > InActive in live data source (beyond Dec 2019)
# > Could exist in JDE/Emerge > needs to be adjusted to inactive
#
# Bucket 4
# > Inactive apple devices
